from django.http import HttpResponse
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from reportlab.pdfgen import canvas
from io import BytesIO
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework import viewsets, permissions, serializers
from rest_framework.response import Response
from rest_framework.views import APIView
from django.utils import timezone
from django.utils.text import slugify
from django.db.models import Q
from django.conf import settings
from apps.events.models import Event
from apps.automation.utils import run_automation
from apps.timeline.models import TimelineEntry
from apps.chats.models import ChatRoom, Message
from .models import (
    PosterDraft,
    Template,
    CertificateRecord,
    CertificateTemplate,
    CertificateTemplateVersion,
    CertificateBatch,
    CertificateBatchItem,
)
from .serializers import (
    CertificateRecordSerializer,
    CertificateTemplateSerializer,
    CertificateTemplateCreateSerializer,
    CertificateBatchSerializer,
    CertificateBatchSubmitSerializer,
)
from .services.template_renderer import render_certificate_html
from .tasks import generate_certificates_batch
import csv, zipfile, io
from PIL import Image
from openpyxl import load_workbook


def _parse_recipients_upload(upload):
    """Extract recipient names from a CSV or XLSX upload."""
    filename = (upload.name or '').lower()
    names = []
    if filename.endswith('.xlsx'):
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_col=1):
            cell = row[0]
            if cell.value:
                names.append(str(cell.value).strip())
        upload.seek(0)
    else:
        try:
            text = upload.read().decode('utf-8')
        except UnicodeDecodeError:
            upload.seek(0)
            text = upload.read().decode('latin-1', errors='ignore')
        for row in csv.reader(io.StringIO(text)):
            if row and row[0].strip():
                names.append(row[0].strip())
        upload.seek(0)
    return names


def _post_system_message(event: Event, text: str):
    """Best-effort system message to the event's primary chat room."""
    room = event.chat_rooms.order_by('id').first()
    if not room:
        return
    Message.objects.create(room=room, sender=None, content=text, attachments=[{'type': 'system'}])


def _render_certificate_pdf(name: str, event_name: str) -> bytes:
    """Render a simple certificate PDF for a recipient and event name."""
    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    p.setFont('Helvetica-Bold', 24)
    p.drawString(100, 750, 'Certificate of Participation')
    p.setFont('Helvetica', 16)
    p.drawString(100, 700, f"This certifies that {name}")
    p.setFont('Helvetica', 12)
    p.drawString(100, 670, f"For: {event_name}")
    p.drawString(100, 640, f"Issued on: {timezone.now().date()}")
    p.showPage(); p.save(); buffer.seek(0)
    return buffer.getvalue()


def _save_certificate_file(content: bytes, path: str) -> tuple[str, str]:
    """Persist a certificate PDF and return storage path and URL."""
    saved_path = default_storage.save(path, ContentFile(content))
    return saved_path, default_storage.url(saved_path)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def certificate_preview(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    p.setFont('Helvetica-Bold', 24)
    p.drawString(100, 750, 'Certificate of Participation')
    p.setFont('Helvetica', 14)
    p.drawString(100, 700, 'This certifies that __________ attended Planora Event')
    p.showPage()
    p.save()
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Cache-Control'] = 'no-store'
    return resp


class PosterDraftViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        u = self.request.user
        return PosterDraft.objects.filter(
            Q(owner=u) |
            Q(room__memberships__user=u) |
            Q(event__owner=u)
        ).distinct().order_by('-updated_at')

    def get_serializer(self, *args, **kwargs):
        class S(serializers.ModelSerializer):
            class Meta:
                model = PosterDraft
                fields = ['id', 'name', 'event', 'room', 'state', 'locked_by', 'locked_at', 'updated_at', 'created_at']
                read_only_fields = ['id', 'locked_by', 'locked_at', 'updated_at', 'created_at']
        return S(*args, **kwargs)

    def perform_create(self, serializer):
        # Enforce user is owner; if room/event provided, ensure permission
        room = serializer.validated_data.get('room')
        event = serializer.validated_data.get('event')
        u = self.request.user
        if room and not (room.event.owner_id == u.id or room.memberships.filter(user=u).exists()):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Forbidden')
        if event and not (event.owner_id == u.id or event.chat_rooms.filter(memberships__user=u).exists()):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Forbidden')
        serializer.save(owner=self.request.user)

    @action(detail=True, methods=['post'])
    def lock(self, request, pk=None):
        draft = self.get_object()
        # Only owner or event owner can lock; members can request lock if unlocked
        u = request.user
        if not (draft.owner_id == u.id or (draft.event and draft.event.owner_id == u.id) or (draft.room and draft.room.memberships.filter(user=u).exists())):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Forbidden')
        if draft.locked_by_id and draft.locked_by_id != u.id:
            return Response({'detail': 'already locked', 'locked_by': draft.locked_by_id}, status=409)
        draft.locked_by = u
        draft.locked_at = timezone.now()
        draft.save(update_fields=['locked_by', 'locked_at'])
        return Response({'status':'locked', 'locked_by': u.id, 'locked_at': draft.locked_at})

    @action(detail=True, methods=['post'])
    def unlock(self, request, pk=None):
        draft = self.get_object()
        u = request.user
        if not (draft.owner_id == u.id or (draft.event and draft.event.owner_id == u.id) or (draft.locked_by_id == u.id)):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Forbidden')
        draft.locked_by = None
        draft.locked_at = None
        draft.save(update_fields=['locked_by', 'locked_at'])
        return Response({'status':'unlocked'})

    @action(detail=True, methods=['post', 'get'])
    def export(self, request, pk=None):
        draft = self.get_object()
        fmt = (request.query_params.get('format') or 'png').lower()

        # Normalize state fields coming from the editor
        state = draft.state or {}
        size = state.get('size') or {}
        try:
            width = int(size.get('w') or size.get('width') or state.get('width') or 800)
        except Exception:
            width = 800
        try:
            height = int(size.get('h') or size.get('height') or state.get('height') or 600)
        except Exception:
            height = 600

        # Extract a safe background color string (hex only)
        def _bg_color():
            raw = state.get('bg') or {}
            if isinstance(raw, str):
                return raw
            if isinstance(raw, dict):
                c = raw.get('color') or raw.get('fill')
                if isinstance(c, str):
                    return c
            return '#ffffff'

        bg = _bg_color()

        if fmt in ('png', 'jpg', 'jpeg'):
            mode = 'RGB'
            # PIL expects either a color tuple or hex; guard against bad values
            try:
                img = Image.new(mode, (width, height), bg)
            except Exception:
                img = Image.new(mode, (width, height), '#ffffff')

            # Note: we don't re-render objects server-side in MVP
            buf = BytesIO()
            img.save(buf, format='PNG' if fmt == 'png' else 'JPEG')
            buf.seek(0)
            resp = HttpResponse(buf.getvalue(), content_type=f'image/{"png" if fmt=="png" else "jpeg"}')
            resp['Content-Disposition'] = f'attachment; filename="poster.{fmt if fmt!="jpeg" else "jpg"}"'
            return resp

        if fmt == 'pdf':
            # Basic PDF with background color; actual scene rendered client-side
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=(width, height))

            # convert hex to 0-1 RGB; fallback to white
            def _hex_to_rgb01(val: str):
                v = (val or '').lstrip('#')
                if len(v) == 3:
                    v = ''.join(ch*2 for ch in v)
                try:
                    r = int(v[0:2], 16) / 255.0
                    g = int(v[2:4], 16) / 255.0
                    b = int(v[4:6], 16) / 255.0
                    return r, g, b
                except Exception:
                    return 1, 1, 1

            r, g, b = _hex_to_rgb01(bg)
            p.setFillColorRGB(r, g, b)
            p.rect(0, 0, width, height, stroke=0, fill=1)
            p.showPage()
            p.save()
            buffer.seek(0)
            resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            resp['Content-Disposition'] = 'attachment; filename="poster.pdf"'
            return resp

        return Response({'detail': 'unsupported format'}, status=400)


class TemplateViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        u = self.request.user
        return Template.objects.filter(Q(owner=u) | Q(owner__isnull=True)).order_by('name')

    class S(serializers.ModelSerializer):
        class Meta:
            model = Template
            fields = ['id', 'name', 'json']

    def get_serializer(self, *args, **kwargs):
        return self.S(*args, **kwargs)

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)


class CertificateRecordViewSet(viewsets.ReadOnlyModelViewSet):
    """List certificate records for event owners and recipients."""

    serializer_class = CertificateRecordSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        u = self.request.user
        qs = CertificateRecord.objects.filter(
            Q(user=u) | Q(event__owner=u) | Q(event__participants=u)
        ).distinct().order_by('-issued_at')
        event_id = self.request.query_params.get('event_id') or self.request.query_params.get('event')
        if event_id:
            qs = qs.filter(event_id=event_id)
        return qs


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def certificates_generate(request):
    """Upload CSV/XLSX of names and return a ZIP of PDFs when the event is LIVE."""
    upload = request.FILES.get('file')
    event_id = request.data.get('event_id') or request.data.get('event') or request.query_params.get('event_id') or request.query_params.get('event')
    if not upload:
        return Response({'detail': 'file required'}, status=400)
    if not event_id:
        return Response({'detail': 'event is required'}, status=400)

    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return Response({'detail': 'event not found'}, status=404)

    if event.owner_id != request.user.id:
        return Response({'detail': 'Only the event owner can generate certificates'}, status=403)
    if (getattr(event, 'status', 'DRAFT') or 'DRAFT').upper() != 'LIVE':
        return Response({'detail': 'Certificates can be generated only when event status is LIVE'}, status=400)

    names = _parse_recipients_upload(upload)
    if not names:
        return Response({'detail': 'No recipient names found'}, status=400)

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name in names:
            pdf_bytes = _render_certificate_pdf(name, event.name)
            safe_name = slugify(name) or 'recipient'
            rel_path = f"certificates/{event.id}/{safe_name}.pdf"
            saved_path, saved_url = _save_certificate_file(pdf_bytes, rel_path)
            CertificateRecord.objects.create(
                user=request.user,
                event=event,
                name=name,
                file=saved_path,
                file_url=saved_url,
            )
            zf.writestr(f"{safe_name}.pdf", pdf_bytes)
    zip_buffer.seek(0)
    run_automation('certificates_generated', {'event_id': event.id, 'count': len(names)})
    _post_system_message(event, f"Certificates generated for {len(names)} participant(s).")
    resp = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    resp['Content-Disposition'] = 'attachment; filename="certificates.zip"'
    return resp


class CertificateTemplateCreateAPIView(APIView):
    """Create a certificate template and its first (or next) version."""

    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        serializer = CertificateTemplateCreateSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        template = data.get('template')
        created = False
        if not template:
            template = CertificateTemplate.objects.create(
                name=data['name'],
                event=data.get('event'),
                is_default=data.get('is_default', False),
                created_by=request.user,
            )
            created = True
        else:
            # Update name/default if provided when adding a new version
            template.name = data.get('name', template.name)
            template.is_default = data.get('is_default', template.is_default)
            template.save(update_fields=['name', 'is_default'])

        assets = {}
        bg_file = request.FILES.get('background') or request.FILES.get('background_file')
        font_file = request.FILES.get('font') or request.FILES.get('font_file')

        if bg_file:
            bg_path = default_storage.save(f"certificate_templates/{template.id}/background/{bg_file.name}", bg_file)
            assets['background'] = bg_path
            assets['background_url'] = default_storage.url(bg_path)
        if font_file:
            font_path = default_storage.save(f"certificate_templates/{template.id}/fonts/{font_file.name}", font_file)
            assets['font'] = font_path
            assets['font_url'] = default_storage.url(font_path)

        version = CertificateTemplateVersion.objects.create(
            template=template,
            html=data.get('html', ''),
            css=data.get('css', ''),
            assets=assets,
            is_published=True,
        )

        template.active_version = version
        template.is_default = data.get('is_default', template.is_default)
        template.save(update_fields=['active_version', 'is_default'])

        run_automation('certificate_template_published', {'template_id': template.id})

        resp_data = CertificateTemplateSerializer(template, context={'request': request}).data
        status_code = 201 if created else 200
        return Response(resp_data, status=status_code)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def certificate_template_preview(request, version_id: int):
    """Render a preview PDF for a published template version."""
    try:
        version = CertificateTemplateVersion.objects.select_related('template', 'template__event').get(id=version_id)
    except CertificateTemplateVersion.DoesNotExist:
        return Response({'detail': 'Template version not found'}, status=404)

    template = version.template
    user = request.user
    # Permission: event participants/owner or staff for event templates; admins only for global templates
    if template.event_id:
        event = template.event
        is_participant = event.participants.filter(id=user.id).exists()
        if not (event.owner_id == user.id or is_participant or user.is_staff):
            return Response({'detail': 'Forbidden'}, status=403)
    else:
        if not user.is_staff:
            return Response({'detail': 'Forbidden'}, status=403)

    payload = request.data if isinstance(request.data, dict) else {}
    try:
        pdf_bytes = render_certificate_html(version, payload)
    except Exception as exc:
        return Response({'detail': f'Render failed: {exc}'}, status=400)

    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="certificate-preview.pdf"'
    return resp


def _log_timeline(event_id, type_, source, payload, actor=None):
    if not event_id:
        return
    payload = payload or {}
    payload.setdefault('snapshot', {})
    TimelineEntry.objects.create(event_id=event_id, actor=actor, type=type_, source=source, payload=payload)


class CertificateBatchSubmitAPIView(APIView):
    """Submit a batch certificate generation job (v2, async)."""

    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'file required'}, status=400)

        serializer = CertificateBatchSubmitSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        event = serializer.validated_data['event']
        template_version = serializer.validated_data['template_version']

        names = _parse_recipients_upload(upload)
        if not names:
            return Response({'detail': 'No recipient names found'}, status=400)

        max_batch = getattr(settings, 'CERTIFICATE_BATCH_MAX', 500)
        if len(names) > max_batch:
            return Response({'detail': f'max {max_batch} recipients per batch'}, status=400)

        batch = CertificateBatch.objects.create(
            event=event,
            template_version=template_version,
            total_count=len(names),
            created_by=request.user,
        )
        items = [CertificateBatchItem(batch=batch, name=n) for n in names]
        CertificateBatchItem.objects.bulk_create(items)

        payload = {'batch_id': str(batch.id), 'event_id': event.id, 'count': len(names)}
        run_automation('certificates_batch_submitted', payload)
        _log_timeline(event.id, 'certificate_batch_submitted', 'certificate', payload, actor=request.user)

        generate_certificates_batch.delay(str(batch.id))

        resp_data = CertificateBatchSerializer(batch, context={'request': request}).data
        return Response(resp_data, status=202)


class CertificateBatchDetailAPIView(APIView):
    """Fetch status of a certificate batch."""

    permission_classes = [IsAuthenticated]

    def get(self, request, batch_id, *args, **kwargs):
        try:
            batch = CertificateBatch.objects.select_related('event', 'template_version', 'created_by').get(id=batch_id)
        except CertificateBatch.DoesNotExist:
            return Response({'detail': 'Batch not found'}, status=404)

        user = request.user
        event = batch.event
        is_participant = event.participants.filter(id=user.id).exists()
        if not (event.owner_id == user.id or is_participant or user.is_staff):
            return Response({'detail': 'Forbidden'}, status=403)

        data = CertificateBatchSerializer(batch, context={'request': request}).data
        return Response(data)
